from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from collections import namedtuple

class StandardExporter:

    def __init__(self, mtd_document, xl_path):
        self.mtd = mtd_document
        self.xl_path = xl_path
        self.workbook = Workbook()

    def export(self):
        for t in self.mtd.tables:

            styles = StandardStyles
            layout = StandardLayout(t)
            content = StandardContent(t)

            ws = self.workbook.create_sheet()
            writer = WorksheetWriter(ws, t, layout, content, styles)
            writer.write()
            writer.merge_cells()
            writer.format()

        # remove first (default) sheet
        first_sheet = self.workbook['Sheet']
        self.workbook.remove(first_sheet)

    def save(self):
        self.workbook.save(self.xl_path)

class WorksheetWriter:

    def __init__(self, worksheet, table, layout, content, styles):
        self.worksheet = worksheet
        self.table = table
        self.layout = layout
        self.content = content
        self.styles = styles

        self.worksheet.title = f'T{table.index}'

        # creates ranges with content dictionary
        self.ranges_with_content = {}
        self.ranges_with_content[layout.top_annotation] = content.top_annotation
        self.ranges_with_content[layout.back_to_content] = content.back_to_content
        self.ranges_with_content[layout.top_banner] = content.top_banner
        self.ranges_with_content[layout.side_banner] = content.side_banner
        self.ranges_with_content[layout.data] = content.data
        self.ranges_with_content[layout.bottom_annotation] = content.bottom_annotation
        self.ranges_with_content.pop(None, None)

    def write(self):

        # calculates worksheet dimensions
        height = max(r.x2 for r in self.ranges_with_content.keys())
        width = max(r.y2 for r in self.ranges_with_content.keys())

        # creates empty cell matrix for the worksheet
        worksheet_cells = [[None for _ in range(width)] for _ in range(height)]

        # populates various ranges with corresponding content
        for r, content in self.ranges_with_content.items():
            for i, row in enumerate(range(r.x1 - 1, r.x2)):
                for j, col in enumerate(range(r.y1 - 1, r.y2)):
                    worksheet_cells[row][col] = content[i][j]

        # remove trailing empty cells and appends cells to the worksheet
        for row in worksheet_cells:
            while row and row[-1] is None:
                row.pop()
            self.worksheet.append(row)

    def merge_cells(self):
        if self.table.top_banner:
            mc = self._get_merged_cells(self.table.top_banner)
            for c in mc:
                #offsetting coordinates
                r = Range(
                    x1=c.x1+self.layout.top_banner.x1,
                    y1=c.y1+self.layout.top_banner.y1,
                    x2=c.x2+self.layout.top_banner.x1,
                    y2=c.y2+self.layout.top_banner.y1)
                self.worksheet.merge_cells(start_row=r.x1, start_column=r.y1, end_row=r.x2, end_column=r.y2)
        if self.table.side_banner:
            mc = self._get_merged_cells(self.table.side_banner)
            for c in mc:
                #offsetting coordinates
                r = Range(
                    x1=c.x1+self.layout.side_banner.x1,
                    y1=c.y1+self.layout.side_banner.y1,
                    x2=c.x2+self.layout.side_banner.x1,
                    y2=c.y2+self.layout.side_banner.y1)
                self.worksheet.merge_cells(start_row=r.x1, start_column=r.y1, end_row=r.x2, end_column=r.y2)

    def format(self):

        if self.layout.top_annotation:
            self._format_annotation(self.layout.top_annotation)
        if self.layout.top_banner:
            self._format_banner(self.layout.top_banner, self.table.top_banner)
        if self.layout.side_banner:
            self._format_banner(self.layout.side_banner, self.table.side_banner)
        if self.layout.data:
            self._format_data()
        if self.layout.bottom_annotation:
            self._format_annotation(self.layout.bottom_annotation)

    def _format_annotation(self, _range):
        for row in range(_range.x1, _range.x2 + 1):
            for col in range(_range.y1, _range.y2 + 1):
                cell = self.worksheet.cell(column=col, row=row)
                style = self.styles.annotation()
                style.apply(cell)

    def _format_banner(self, _range, banner):
        for i, row in enumerate(range(_range.x1, _range.x2 + 1)):
            for j, col in enumerate(range(_range.y1, _range.y2 + 1)):
                cell = self.worksheet.cell(column=col, row=row)
                banner_cell = banner.banner[i][j]
                if banner_cell.type == 'Axis':
                    style = self.styles.banner_axis(banner_cell)
                    style.apply(cell)
                elif banner_cell.type == 'Element':
                    style = self.styles.banner_element(banner_cell)
                    style.apply(cell)
                elif banner_cell.type == '':
                    style = self.styles.banner_empty(banner_cell)
                    style.apply(cell)

    def _format_data(self):

        r = self.layout.data
        top_base_mask = self.table.top_banner.base_mask if self.table.top_banner else [False] * (r.y2 - r.y1 + 1)
        top_first_mask = self.table.top_banner.first_mask if self.table.top_banner else [False] * (r.y2 - r.y1 + 1)
        side_base_mask = self.table.side_banner.base_mask if self.table.side_banner else [False] * (r.x2 - r.x1 + 1)
        side_first_mask = self.table.side_banner.first_mask if self.table.side_banner else [False] * (r.x2 - r.x1 + 1)
        side_cell_items_mask = self.table.side_banner.cell_items_mask if self.table.side_banner else [False] * (r.x2 - r.x1 + 1)
        side_last_element_mask = self.table.side_banner.last_element_mask if self.table.side_banner else [False] * (r.x2 - r.x1 + 1)
        for i, row in enumerate(range(r.x1, r.x2 + 1)):
            for j, col in enumerate(range(r.y1, r.y2 + 1)):
                cell = self.worksheet.cell(column=col, row=row)
                is_base = any((side_base_mask[i], top_base_mask[j]))
                is_top_first = top_first_mask[j]
                is_side_first = side_first_mask[i]
                cell_item = side_cell_items_mask[i]
                last_element = side_last_element_mask[i]
                show_perc = self.table.show_perc_signs
                style = self.styles.data_cell(cell, is_base, is_top_first, is_side_first, cell_item, last_element, show_perc)
                style.apply(cell)

    def _get_merged_cells(self, banner):
        
        merged_cells = []

        if banner.name == 'Top':
            banner._transpose()

        # merging empty cells to existing (horizontal)
        for i, row in enumerate(banner.banner):
            for j, cell in enumerate(row):
                if not cell.object:
                    merged_cells.append(Range(i, j - 1, i, banner.width - 1))
                    break

        # merging due to stretched axis/elements (vertical)
        merge_pending = False
        x1, y1, x2, y2 = 0, 0, 0, 0
        for col in range(banner.width):
            for row in range(banner.height): # starts looking from 2nd row
                current_cell = banner.banner[row][col]
                last_cell = banner.banner[row - 1][col] if row > 0 else None
                neighbor_matched = bool(last_cell and last_cell.object and last_cell.object is current_cell.object)
                if not merge_pending and neighbor_matched:
                    x1, y1 = row - 1, col
                    merge_pending = True
                elif merge_pending and not neighbor_matched:
                    x2, y2 = row - 1, col
                    merged_cells.append(Range(x1, y2, x2, y2))
                    merge_pending = False
            if merge_pending:
                x2, y2 = banner.height - 1, col
                merged_cells.append(Range(x1, y1, x2, y2))
                merge_pending = False

        # transposes if top banner
        if banner.name == 'Top':
            banner._transpose()
            merged_cells = [Range(c.y1, c.x1, c.y2, c.x2)
                    for c in merged_cells]

        return merged_cells

class StandardLayout:

    def __init__(self, table):

        self.table = table
        current_row = 1

        # top annotation
        annotations = [a for a in self.table.top_annotations if a]
        if annotations:
            self.top_annotation = Range(current_row, 1, len(annotations), 1)
            current_row = self.top_annotation.x2 + 2
        else:
            self.top_annotation = None

        # back to content
        self.back_to_content = None
        
        # top banner
        if self.table.top_banner:
            x1 = current_row
            y1 = self.table.side_banner.width + 1 if self.table.side_banner else 0
            x2 = x1 + self.table.top_banner.height - 1
            y2 = y1 + self.table.top_banner.width - 1
            self.top_banner = Range(x1, y1, x2, y2)
            current_row = self.top_banner.x2 + 1
        else:
            self.top_banner = None

        # side banner
        if self.table.side_banner:
            x1 = current_row
            y1 = 1
            x2 = x1 + self.table.side_banner.height - 1
            y2 = y1 + self.table.side_banner.width - 1
            self.side_banner = Range(x1, y1, x2, y2)
            current_row = self.side_banner.x2 + 1
        else:
            self.side_banner = None

        # data
        if self.table.data:
            x1 = self.side_banner.x1 if self.side_banner else current_row
            y1 = self.side_banner.y2 + 1 if self.side_banner else 0
            x2 = x1 + len(self.table.data) - 1
            y2 = y1 + len(self.table.data[0]) - 1
            self.data = Range(x1, y1, x2, y2)
            current_row = self.data.x2 + 2
        else:
            self.data = None

        # bottom annotations
        annotations = [a for a in self.table.bottom_annotations if a]
        if annotations:
            self.bottom_annotation = Range(current_row, 1, current_row + len(annotations) - 1, 1)
            current_row = self.bottom_annotation.x2 + 1
        else:
            self.bottom_annotation = None

class StandardContent:
     
    def __init__(self, table):
         
        self.table = table
        self.top_annotation = [[a] for a in self.table.top_annotations if a] if table.top_annotations else None
        self.back_to_content = None
        self.top_banner = [[cell.label for cell in row] for row in self.table.top_banner.banner] if table.top_banner else None
        self.side_banner = [[cell.label for cell in row] for row in self.table.side_banner.banner] if table.side_banner else None
        self.data = self.table.data if self.table.data else None
        self.bottom_annotation = [[a] for a in self.table.bottom_annotations if a] if table.bottom_annotations else None

class StandardStyles:

    @staticmethod
    def annotation():
        font = Font(name='Arial', sz=12)
        return CellStyle(font=font)

    @staticmethod
    def banner_element(cell):
        font_size = 8 - cell.element_level / 2
        italic = bool(cell.element_level % 2)
        bold = 'Base' in cell.object.type if cell.object else False
        indent = cell.element_level
        font = Font(name='Arial', sz=font_size, italic=italic, bold=bold)
        
        a = 'center' if cell.banner.name == 'Top' else 'left'
        alignment = Alignment(horizontal=a, vertical='center', indent=indent, wrap_text=True)

        number_format = '@'
        
        if cell.first:
            if cell.banner.name == 'Top':
                border = Border(left=Side(style='thin', color='FF000000'))
            else:
                border = Border(top=Side(style='thin', color='FF000000'))
            return CellStyle(font=font, alignment=alignment, border=border, number_format=number_format)
        else:
            return CellStyle(font=font, alignment=alignment, number_format=number_format)

    @staticmethod
    def banner_axis(cell):
        font = Font(name='Arial', sz=10, bold=True)
        number_format = '@'
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if cell.banner.name == 'Side':
                alignment.text_rotation=90

        if cell.first:
            if cell.banner.name == 'Top':
                border = Border(left=Side(style='thin', color='FF000000'))
            else:
                border = Border(top=Side(style='thin', color='FF000000'))
            return CellStyle(font=font, alignment=alignment, border=border, number_format=number_format)
        else:
            return CellStyle(font=font, alignment=alignment, number_format=number_format)
        
    @staticmethod
    def banner_empty(cell):

        number_format = '@'        
        if cell.first:
            if cell.banner.name == 'Top':
                border = Border(left=Side(style='thin', color='FF000000'))
            else:
                border = Border(top=Side(style='thin', color='FF000000'))
            return CellStyle(border=border, number_format=number_format)
        else:
            return CellStyle(number_format=number_format)
    
    @staticmethod   
    def data_cell(cell, is_base, is_top_first, is_side_first, cell_item, last_element, show_perc):

        # font
        font = Font(name='Arial', sz=8)
        if is_base:
            font.bold = True
        if 'Percent' in cell_item.type and not is_base:
            font.italic = True

        # border
        border = None
        if is_top_first and is_side_first:
            border=Border(top=Side(style='thin', color='FF000000'),
                left=Side(style='thin', color='FF000000'))
        elif is_top_first and not is_side_first:
            border=Border(left=Side(style='thin', color='FF000000'))
        elif not is_top_first and is_side_first:
            border=Border(top=Side(style='thin', color='FF000000'))

        # number format
        number_format = None

        # takes number of decimal from last element if > 0
        # otherwise from cell_items
        decimals = last_element.decimals or cell_item.decimals

        if decimals > 0:
            number_format = f'0.{"0" * decimals}'
        else:
            number_format = '0'
        if show_perc and 'Percent' in cell_item.type:
            number_format += '%'

        return CellStyle(font=font, border=border,number_format=number_format)

class CellStyle:

    def __init__(self, font=None, fill=None, border=None, alignment=None, number_format=None):
        self.font = font
        self.fill = fill
        self.border = border
        self.alignment = alignment
        self.number_format = number_format

    def apply(self, cell):
        if self.font:
            cell.font = self.font
        if self.fill:
            cell.fill = self.fill
        if self.border:
            cell.border = self.border
        if self.alignment:
            cell.alignment = self.alignment
        if self.number_format:
            cell.number_format = self.number_format

Range = namedtuple('Range', 'x1 y1 x2 y2')